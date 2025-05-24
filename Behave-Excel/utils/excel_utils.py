from openpyxl import load_workbook

class ExcelUtil:
    def __init__(self, path, sheet_name):
        # passing the path and sheet name
        self.path = path
        # creating workbook object with path
        self.workbook = load_workbook(path)
        # creating sheet from workbook with sheet name
        self.sheet = self.workbook[sheet_name]

    def read_data(self, row_no, col_no):
        # return the exact data present in the cell
        return self.sheet.cell(row=row_no, column=col_no).value

    def write_data(self, row_no, col_no, data):
        # write the data in exact cell with data passed
        self.sheet.cell(row=row_no, column=col_no).value = data

    def row_count(self):
        # max rows present in excel sheet
        return self.sheet.max_row
